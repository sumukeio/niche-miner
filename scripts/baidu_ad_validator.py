"""
百度竞价关键词商业价值验证工具
用于批量检测关键词是否有百度广告投放，验证商业价值
"""

import os
import time
import random
import pandas as pd
from pathlib import Path
from playwright.sync_api import sync_playwright, Page, TimeoutError as PlaywrightTimeoutError
from typing import List, Dict, Optional, Tuple
import logging

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
    ]
)
logger = logging.getLogger(__name__)


class BaiduAdValidator:
    """百度广告验证器"""
    
    def __init__(self, headless: bool = False, screenshot_dir: str = "scripts/screenshots", 
                 proxy: Optional[str] = None, mobile: bool = False, proxy_list: Optional[List[str]] = None):
        """
        初始化验证器
        
        Args:
            headless: 是否无头模式运行（False 表示显示浏览器窗口）
            screenshot_dir: 截图保存目录
            proxy: 代理服务器地址（格式：http://host:port 或 socks5://host:port）
            mobile: 是否使用移动端模式
            proxy_list: 代理IP列表（从文件读取，自动轮换）
        """
        self.headless = headless
        self.screenshot_dir = Path(screenshot_dir)
        self.screenshot_dir.mkdir(exist_ok=True)
        self.mobile = mobile
        self.proxy = proxy
        self.proxy_list = proxy_list or []
        self.proxy_index = 0  # 当前使用的代理索引
        
        # 根据模式选择 User-Agent
        if mobile:
            # 移动端 User-Agent (iPhone)
            self.user_agent = (
                "Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) "
                "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1"
            )
            self.viewport = {'width': 390, 'height': 844}  # iPhone 14 Pro 尺寸
            self.is_mobile = True
        else:
            # PC 端 User-Agent
            self.user_agent = (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            )
            self.viewport = {'width': 1920, 'height': 1080}
            self.is_mobile = False
        
        # 需要过滤的平台域名列表
        # 注意：只需要主域名，子域名会自动匹配（如 item.jd.com 会匹配 jd.com）
        self.filtered_domains = [
            'jd.com', 'jd.hk',                    # 京东
            '1688.com', 'alibaba.com', 'alibaba.com.cn',  # 阿里巴巴/1688
            'b2b.baidu.com', 'aicaigou.com',      # 爱采购
            'zhihu.com',                          # 知乎
            'tmall.com', 'tmall.hk'               # 天猫
        ]
    
    def get_current_proxy(self) -> Optional[Dict[str, str]]:
        """
        获取当前使用的代理配置
        
        Returns:
            代理配置字典，格式：{'server': 'http://host:port'} 或 None
        """
        proxy_to_use = None
        
        # 优先使用代理列表（轮换）
        if self.proxy_list:
            if self.proxy_index < len(self.proxy_list):
                proxy_to_use = self.proxy_list[self.proxy_index]
                # 下次使用下一个代理（轮换）
                self.proxy_index = (self.proxy_index + 1) % len(self.proxy_list)
                logger.info(f"使用代理: {proxy_to_use} (索引: {self.proxy_index - 1 if self.proxy_index > 0 else len(self.proxy_list) - 1}/{len(self.proxy_list)})")
        # 其次使用单个代理
        elif self.proxy:
            proxy_to_use = self.proxy
        
        if proxy_to_use:
            # 确保代理格式正确
            if not proxy_to_use.startswith(('http://', 'https://', 'socks5://')):
                proxy_to_use = 'http://' + proxy_to_use
            return {'server': proxy_to_use}
        
        return None
    
    def wait_random(self, min_seconds: float = 2.0, max_seconds: float = 5.0):
        """
        随机等待，模拟真人操作
        
        Args:
            min_seconds: 最小等待时间（秒）
            max_seconds: 最大等待时间（秒）
        """
        wait_time = random.uniform(min_seconds, max_seconds)
        time.sleep(wait_time)
    
    def extract_domain(self, url: str) -> str:
        """
        从 URL 中提取域名（支持百度跳转链接）
        
        Args:
            url: 完整的 URL 字符串
            
        Returns:
            域名（如：jd.com）
        """
        try:
            from urllib.parse import urlparse, parse_qs, unquote
            real_url = url
            
            # 处理百度跳转链接 (https://www.baidu.com/link?url=xxx)
            if 'link?url=' in url or '/link?url=' in url:
                try:
                    parsed = urlparse(url)
                    query_params = parse_qs(parsed.query)
                    if 'url' in query_params:
                        # 获取真实链接（可能需要URL解码）
                        real_url = unquote(query_params['url'][0])
                    elif 'url=' in url:
                        # 备用方法：直接从字符串中提取
                        parts = url.split('url=')
                        if len(parts) > 1:
                            real_url = unquote(parts[1].split('&')[0])
                except Exception:
                    # 如果解析失败，尝试简单字符串提取
                    if 'url=' in url:
                        try:
                            real_url = unquote(url.split('url=')[1].split('&')[0])
                        except:
                            pass
            
            # 解析真实URL的域名
            parsed = urlparse(real_url)
            domain = parsed.netloc.lower()
            
            # 移除 www. 前缀
            if domain.startswith('www.'):
                domain = domain[4:]
            
            # 处理端口号（如 jd.com:443）
            if ':' in domain:
                domain = domain.split(':')[0]
            
            return domain
        except Exception as e:
            logger.debug(f"提取域名失败: {url} - {str(e)}")
            return ""
    
    def is_filtered_platform(self, url: str) -> bool:
        """
        判断链接是否来自需要过滤的平台
        
        Args:
            url: 广告链接
            
        Returns:
            True 如果是需要过滤的平台，False 否则
        """
        domain = self.extract_domain(url)
        if not domain:
            # 如果提取域名失败，尝试直接从 URL 中检查（备用方案）
            url_lower = url.lower()
            for filtered in self.filtered_domains:
                if filtered in url_lower:
                    logger.debug(f"通过URL字符串匹配过滤: {filtered} in {url}")
                    return True
            return False
        
        # 检查是否匹配任何过滤域名
        # 支持精确匹配和子域名匹配（如 item.jd.com 匹配 jd.com）
        for filtered in self.filtered_domains:
            if domain == filtered:
                logger.debug(f"精确匹配过滤域名: {domain} == {filtered}")
                return True
            elif domain.endswith('.' + filtered):
                logger.debug(f"子域名匹配过滤域名: {domain} ends with .{filtered}")
                return True
        
        return False
    
    def search_keyword(self, page: Page, keyword: str, use_direct_url: bool = False) -> bool:
        """
        在百度搜索关键词
        
        Args:
            page: Playwright Page 对象
            keyword: 要搜索的关键词
            use_direct_url: 是否直接使用 URL 参数搜索（备选方案）
            
        Returns:
            是否成功加载搜索结果页
        """
        try:
            if use_direct_url:
                # 方法2: 直接使用 URL 参数搜索（更稳定，避免交互问题）
                from urllib.parse import quote
                encoded_keyword = quote(keyword)
                # 根据模式选择百度搜索URL
                if self.is_mobile:
                    search_url = f"https://m.baidu.com/s?wd={encoded_keyword}"
                else:
                    search_url = f"https://www.baidu.com/s?wd={encoded_keyword}"
                page.goto(search_url, timeout=30000, wait_until="domcontentloaded")
                page.wait_for_timeout(2000)
            else:
                # 方法1: 通过首页搜索框搜索（更接近真实用户行为）
                # 访问百度首页（根据模式选择PC端或移动端）
                if self.is_mobile:
                    baidu_url = "https://m.baidu.com"
                else:
                    baidu_url = "https://www.baidu.com"
                page.goto(baidu_url, timeout=30000, wait_until="domcontentloaded")
                
                # 等待页面加载
                page.wait_for_timeout(2000)
                
                # 尝试定位搜索框（优先检查可见性）
                try:
                    # 等待搜索框可见（超时时间缩短到5秒，快速失败）
                    search_input = page.wait_for_selector("#kw", state="visible", timeout=5000)
                    logger.debug(f"搜索框可见，使用正常方式搜索: {keyword}")
                except PlaywrightTimeoutError:
                    # 如果搜索框不可见，立即切换到 URL 方式，避免后续 fill() 超时
                    logger.info(f"搜索框不可见，改用直接 URL 方式: {keyword}")
                    return self.search_keyword(page, keyword, use_direct_url=True)
                
                # 如果搜索框可见，正常输入和搜索
                try:
                    # 模拟鼠标移动到搜索框
                    box = search_input.bounding_box()
                    if box:
                        page.mouse.move(box['x'] + box['width']/2, box['y'] + box['height']/2)
                        page.wait_for_timeout(200)
                    
                    # 点击搜索框使其获得焦点
                    search_input.click(timeout=2000)
                    page.wait_for_timeout(300)  # 增加等待时间，让页面响应
                except:
                    pass
                
                # 输入关键词（使用 type 方法模拟真实打字，而不是直接 fill）
                try:
                    # Playwright 的 type 方法（注意：需要先聚焦元素）
                    search_input.focus()
                    page.keyboard.type(keyword, delay=random.randint(50, 150))  # 模拟打字延迟
                    page.wait_for_timeout(300)
                except:
                    # 如果 type 失败，回退到 fill
                    try:
                        search_input.fill(keyword)
                        page.wait_for_timeout(300)
                    except:
                        # 如果 fill 也失败，使用 JavaScript 直接设置值
                        page.evaluate(f"""
                            const input = document.querySelector('#kw');
                            if (input) {{
                                input.value = '{keyword}';
                                input.dispatchEvent(new Event('input', {{ bubbles: true }}));
                            }}
                        """)
                        page.wait_for_timeout(300)
                
                # 点击搜索按钮或按回车
                search_button = page.query_selector("#su")
                if search_button:
                    try:
                        # 模拟鼠标移动到搜索按钮
                        box = search_button.bounding_box()
                        if box:
                            page.mouse.move(box['x'] + box['width']/2, box['y'] + box['height']/2)
                            page.wait_for_timeout(200)
                        
                        search_button.click()
                    except:
                        # 如果按钮点击失败，尝试按回车
                        search_input.press("Enter")
                else:
                    search_input.press("Enter")
            
            # 等待搜索结果页加载
            try:
                # 等待搜索结果容器出现
                page.wait_for_selector("#content_left, .s_main, .result", timeout=20000)
            except PlaywrightTimeoutError:
                # 如果容器未出现，检查 URL 是否已跳转
                page.wait_for_timeout(3000)
                if "baidu.com/s" not in page.url and not use_direct_url:
                    logger.warning(f"搜索结果页加载异常，尝试直接 URL 方式: {keyword}")
                    return self.search_keyword(page, keyword, use_direct_url=True)
                elif not is_search_page:
                    logger.error(f"搜索结果页加载失败: {keyword}")
                    return False
            
            # 额外等待，确保广告标签渲染完成
            page.wait_for_timeout(2000)
            
            return True
            
        except PlaywrightTimeoutError as e:
            logger.error(f"搜索超时: {keyword} - {str(e)}")
            # 如果超时，尝试直接 URL 方式
            if not use_direct_url:
                logger.info(f"尝试使用直接 URL 方式重试: {keyword}")
                return self.search_keyword(page, keyword, use_direct_url=True)
            return False
        except Exception as e:
            logger.error(f"搜索异常: {keyword} - {str(e)}")
            # 如果异常，尝试直接 URL 方式
            if not use_direct_url:
                logger.info(f"尝试使用直接 URL 方式重试: {keyword}")
                return self.search_keyword(page, keyword, use_direct_url=True)
            return False
    
    def detect_ads(self, page: Page, keyword: str) -> Tuple[bool, List[Dict[str, str]]]:
        """
        检测搜索结果中的广告
        
        Args:
            page: Playwright Page 对象
            keyword: 当前搜索的关键词（用于截图文件名）
            
        Returns:
            (是否有广告, 广告信息列表)
            广告信息格式: [{"title": "广告标题", "link": "广告链接"}, ...]
        """
        has_ads = False
        ad_info_list = []
        
        try:
            # 方法1: 查找包含"广告"文本的元素
            # 百度的广告通常有一个包含"广告"的 span 或标签
            ad_labels = page.query_selector_all('span:has-text("广告"), .ad-label, .ad-text, [class*="ad"] span:has-text("广告")')
            
            # 方法2: 查找带有特定 class 的广告结果容器
            # 百度广告结果通常有特定的 class，如 "c-container" 且包含 "广告" 标识
            ad_containers = page.query_selector_all('.c-container')
            
            # 更精确的方法：查找结果项，检查其内部是否包含"广告"标识
            result_items = page.query_selector_all('#content_left .c-container, #content_left .result')
            
            found_ad_count = 0
            for item in result_items[:10]:  # 只检查前10个结果
                try:
                    # 获取整个结果项的文本内容
                    item_text = item.inner_text()
                    
                    # 检查是否包含"广告"标识
                    if "广告" in item_text or "推广" in item_text:
                        # 提取广告标题和链接
                        title_elem = item.query_selector('h3 a, .t a, a[href]')
                        link_elem = item.query_selector('a[href]')
                        
                        if title_elem and link_elem:
                            title = title_elem.inner_text().strip()
                            link = link_elem.get_attribute('href') or ""
                            
                            # 处理相对链接
                            if link.startswith('/'):
                                link = f"https://www.baidu.com{link}"
                            
                            # 注意：百度跳转链接会在 is_filtered_platform() 中自动解析和处理
                            
                            if title and link:
                                # 检查是否需要过滤该平台
                                if self.is_filtered_platform(link):
                                    logger.info(f"过滤平台广告: {title[:30]}... - {link[:80]}")
                                    continue  # 跳过该广告
                                
                                ad_info_list.append({
                                    "title": title,
                                    "link": link
                                })
                                found_ad_count += 1
                                
                                if found_ad_count >= 3:  # 只取前3个
                                    break
                                    
                except Exception as e:
                    # 某个结果项解析失败，继续下一个
                    continue
            
            has_ads = len(ad_info_list) > 0
            
            # 如果有广告，保存截图
            if has_ads:
                try:
                    # 清理文件名（移除特殊字符）
                    safe_filename = "".join(c for c in keyword if c.isalnum() or c in (' ', '-', '_')).rstrip()
                    screenshot_path = self.screenshot_dir / f"{safe_filename}.png"
                    
                    # 滚动到页面顶部，确保截图完整
                    page.evaluate("window.scrollTo(0, 0)")
                    page.wait_for_timeout(500)
                    
                    page.screenshot(path=str(screenshot_path), full_page=False)
                    logger.info(f"已保存截图: {screenshot_path}")
                except Exception as e:
                    logger.warning(f"截图保存失败: {keyword} - {str(e)}")
            
        except Exception as e:
            logger.error(f"检测广告时出错: {keyword} - {str(e)}")
        
        return has_ads, ad_info_list
    
    def validate_keyword(self, page: Page, keyword: str, index: int, total: int) -> Dict:
        """
        验证单个关键词
        
        Args:
            page: Playwright Page 对象
            keyword: 关键词
            index: 当前索引（从1开始）
            total: 总数量
            
        Returns:
            验证结果字典
        """
        logger.info(f"[{index}/{total}] 搜索: {keyword}")
        
        # 搜索关键词
        search_success = self.search_keyword(page, keyword)
        
        if not search_success:
            return {
                "keyword": keyword,
                "has_ads": "Error",
                "ad_info_list": []
            }
        
        # 随机等待，模拟真人操作
        self.wait_random(2, 5)
        
        # 检测广告
        has_ads, ad_info_list = self.detect_ads(page, keyword)
        
        # 格式化结果 - 使用列表格式，后续会分开列显示
        # 注意：如果所有广告都被过滤掉（ad_info_list为空），则标记为"No"
        if has_ads and len(ad_info_list) > 0:
            logger.info(f"[{index}/{total}] ✓ {keyword} -> 发现广告！({len(ad_info_list)}个)")
            # 保存为列表格式，方便后续分开列显示
            return {
                "keyword": keyword,
                "has_ads": "Yes",
                "ad_info_list": ad_info_list  # 保存为列表
            }
        else:
            if has_ads and len(ad_info_list) == 0:
                logger.info(f"[{index}/{total}] ✗ {keyword} -> 广告已过滤（全部为平台广告）")
            else:
                logger.info(f"[{index}/{total}] ✗ {keyword} -> 无广告")
            return {
                "keyword": keyword,
                "has_ads": "No",
                "ad_info_list": []  # 空列表
            }
    
    def validate_batch(self, keywords: List[str]) -> List[Dict]:
        """
        批量验证关键词
        
        Args:
            keywords: 关键词列表
            
        Returns:
            验证结果列表
        """
        results = []
        total = len(keywords)
        
        with sync_playwright() as p:
            # 获取当前使用的代理
            current_proxy = self.get_current_proxy()
            
            # 启动浏览器（增强反反爬配置）
            browser = p.chromium.launch(
                headless=self.headless,
                args=[
                    '--disable-blink-features=AutomationControlled',  # 隐藏自动化特征
                    '--disable-dev-shm-usage',
                    '--no-sandbox',
                    '--disable-setuid-sandbox',
                    '--disable-web-security',
                    '--disable-features=IsolateOrigins,site-per-process'
                ],
                proxy=current_proxy  # 设置代理
            )
            
            # 创建新页面（增强反反爬配置）
            context_options = {
                'user_agent': self.user_agent,
                'viewport': self.viewport,
                'locale': 'zh-CN',
                'timezone_id': 'Asia/Shanghai',
                'permissions': ['geolocation'],
                'geolocation': {'latitude': 39.9042, 'longitude': 116.4074},  # 北京坐标
                'color_scheme': 'light'
            }
            
            # 移动端额外配置
            if self.is_mobile:
                context_options['is_mobile'] = True
                context_options['has_touch'] = True
                context_options['device_scale_factor'] = 3
                context_options['screen'] = self.viewport
            
            context = browser.new_context(**context_options)
            
            # 注入 JavaScript 来隐藏自动化特征
            context.add_init_script("""
                // 隐藏 webdriver 特征
                Object.defineProperty(navigator, 'webdriver', {
                    get: () => undefined
                });
                
                // 覆盖 plugins 和 languages
                Object.defineProperty(navigator, 'plugins', {
                    get: () => [1, 2, 3, 4, 5]
                });
                
                Object.defineProperty(navigator, 'languages', {
                    get: () => ['zh-CN', 'zh', 'en']
                });
                
                // 覆盖 permissions
                const originalQuery = window.navigator.permissions.query;
                window.navigator.permissions.query = (parameters) => (
                    parameters.name === 'notifications' ?
                        Promise.resolve({ state: Notification.permission }) :
                        originalQuery(parameters)
                );
                
                // 覆盖 chrome 对象
                window.chrome = {
                    runtime: {}
                };
            """)
            
            page = context.new_page()
            
            try:
                for index, keyword in enumerate(keywords, 1):
                    try:
                        result = self.validate_keyword(page, keyword, index, total)
                        results.append(result)
                    except Exception as e:
                        logger.error(f"处理关键词时出错: {keyword} - {str(e)}")
                        results.append({
                            "keyword": keyword,
                            "has_ads": "Error",
                            "ad_info_list": []
                        })
                    
                    # 每次搜索后随机等待（最后一次不需要）
                    if index < total:
                        self.wait_random(2, 5)
                        
            finally:
                browser.close()
        
        return results


def load_keywords_from_excel(file_path: str = "keywords.xlsx", column_name: str = "Keyword") -> List[str]:
    """
    从 Excel 文件加载关键词
    
    Args:
        file_path: Excel 文件路径
        column_name: 关键词列名
        
    Returns:
        关键词列表
    """
    try:
        df = pd.read_excel(file_path)
        
        if column_name not in df.columns:
            raise ValueError(f"Excel 文件中未找到列 '{column_name}'。可用列: {list(df.columns)}")
        
        # 去除空值和重复
        keywords = df[column_name].dropna().drop_duplicates().tolist()
        
        # 转换为字符串并去除空白
        keywords = [str(k).strip() for k in keywords if str(k).strip()]
        
        logger.info(f"成功加载 {len(keywords)} 个关键词")
        return keywords
        
    except FileNotFoundError:
        logger.error(f"文件未找到: {file_path}")
        raise
    except Exception as e:
        logger.error(f"读取 Excel 文件时出错: {str(e)}")
        raise


def save_results_to_excel(original_file: str, results: List[Dict], output_file: str = "keywords_validated.xlsx"):
    """
    保存验证结果到 Excel 文件
    将每个广告的标题和链接分开列显示，方便查看和点击
    
    Args:
        original_file: 原始 Excel 文件路径
        results: 验证结果列表（包含 ad_info_list）
        output_file: 输出文件路径
    """
    try:
        # 读取原始数据
        original_df = pd.read_excel(original_file)
        
        # 处理结果数据：将广告信息分开列
        processed_results = []
        for result in results:
            processed_result = {
                "keyword": result.get("keyword", ""),
                "Has_Ads": result.get("has_ads", "No")
            }
            
            # 处理广告列表，最多3个广告
            ad_info_list = result.get("ad_info_list", [])
            
            # 为每个广告创建独立的列（Ad_Title_1, Ad_Link_1, Ad_Title_2, Ad_Link_2, Ad_Title_3, Ad_Link_3）
            for i in range(1, 4):  # 最多3个广告
                if i <= len(ad_info_list):
                    ad = ad_info_list[i - 1]
                    processed_result[f"Ad_Title_{i}"] = ad.get("title", "")
                    processed_result[f"Ad_Link_{i}"] = ad.get("link", "")
                else:
                    processed_result[f"Ad_Title_{i}"] = ""
                    processed_result[f"Ad_Link_{i}"] = ""
            
            processed_results.append(processed_result)
        
        # 创建结果 DataFrame
        results_df = pd.DataFrame(processed_results)
        
        # 将结果合并到原始数据
        # 使用关键词作为匹配键
        merged_df = original_df.merge(
            results_df,
            left_on="Keyword",
            right_on="keyword",
            how="left"
        )
        
        # 清理列名（删除重复的 keyword 列）
        if 'keyword' in merged_df.columns and 'Keyword' in merged_df.columns:
            merged_df = merged_df.drop(columns=['keyword'])
        
        # 使用 openpyxl 引擎，支持超链接
        from openpyxl import load_workbook
        from openpyxl.styles import Font
        
        # 先保存为 Excel
        merged_df.to_excel(output_file, index=False, engine='openpyxl')
        
        # 读取工作簿，为链接列添加超链接格式
        wb = load_workbook(output_file)
        ws = wb.active
        
        # 找到所有 Ad_Link_ 列
        header_row = 1
        link_columns = {}
        for col_idx, cell in enumerate(ws[header_row], start=1):
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("Ad_Link_"):
                link_columns[cell.value] = col_idx
        
        # 为链接单元格添加超链接和蓝色字体
        link_font = Font(color="0563C1", underline="single")
        for row_idx in range(2, ws.max_row + 1):
            for col_name, col_idx in link_columns.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value and str(cell.value).strip():
                    url = str(cell.value).strip()
                    # 添加超链接
                    cell.hyperlink = url
                    cell.font = link_font
                    # 将链接文本改为更友好的格式（可选：显示为"点击查看"）
                    # cell.value = "点击查看"  # 如果想隐藏URL，取消注释这行
        
        wb.save(output_file)
        logger.info(f"结果已保存到: {output_file}（链接已格式化为可点击的超链接）")
        
    except Exception as e:
        logger.error(f"保存结果时出错: {str(e)}")
        # 如果合并失败，直接保存结果（不带超链接格式）
        try:
            processed_results = []
            for result in results:
                processed_result = {
                    "keyword": result.get("keyword", ""),
                    "Has_Ads": result.get("has_ads", "No")
                }
                ad_info_list = result.get("ad_info_list", [])
                for i in range(1, 4):
                    if i <= len(ad_info_list):
                        ad = ad_info_list[i - 1]
                        processed_result[f"Ad_Title_{i}"] = ad.get("title", "")
                        processed_result[f"Ad_Link_{i}"] = ad.get("link", "")
                    else:
                        processed_result[f"Ad_Title_{i}"] = ""
                        processed_result[f"Ad_Link_{i}"] = ""
                processed_results.append(processed_result)
            
            results_df = pd.DataFrame(processed_results)
            results_df.to_excel(output_file, index=False)
            logger.info(f"已保存简化结果到: {output_file}")
        except Exception as e2:
            logger.error(f"保存简化结果也失败: {str(e2)}")


def load_proxy_list(file_path: str) -> List[str]:
    """
    从文件加载代理列表（每行一个代理地址）
    
    Args:
        file_path: 代理列表文件路径
        
    Returns:
        代理地址列表
    """
    proxies = []
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):  # 忽略空行和注释
                    proxies.append(line)
        logger.info(f"成功加载 {len(proxies)} 个代理地址")
    except FileNotFoundError:
        logger.error(f"代理列表文件未找到: {file_path}")
    except Exception as e:
        logger.error(f"读取代理列表文件时出错: {str(e)}")
    
    return proxies


def main():
    """主函数"""
    import argparse
    
    parser = argparse.ArgumentParser(description='百度竞价关键词商业价值验证工具')
    parser.add_argument('--input', '-i', default='keywords.xlsx', help='输入 Excel 文件路径 (默认: keywords.xlsx)')
    parser.add_argument('--output', '-o', default='keywords_validated.xlsx', help='输出 Excel 文件路径 (默认: keywords_validated.xlsx)')
    parser.add_argument('--column', '-c', default='Keyword', help='关键词列名 (默认: Keyword)')
    parser.add_argument('--headless', action='store_true', help='无头模式运行（不显示浏览器窗口）')
    parser.add_argument('--screenshots', '-s', default='scripts/screenshots', help='截图保存目录 (默认: scripts/screenshots)')
    
    args = parser.parse_args()
    
    logger.info("=" * 60)
    logger.info("百度竞价关键词商业价值验证工具")
    logger.info("=" * 60)
    
    try:
        # 加载关键词
        keywords = load_keywords_from_excel(args.input, args.column)
        
        if not keywords:
            logger.error("未找到有效关键词，程序退出")
            return
        
        # 加载代理列表（如果指定了文件）
        proxy_list = None
        if args.proxy_list:
            proxy_list = load_proxy_list(args.proxy_list)
            if not proxy_list:
                logger.warning("代理列表为空，将不使用代理")
        
        # 创建验证器
        validator = BaiduAdValidator(
            headless=args.headless,
            screenshot_dir=args.screenshots,
            proxy=args.proxy,
            mobile=args.mobile,
            proxy_list=proxy_list
        )
        
        if args.mobile:
            logger.info("使用移动端模式（模拟手机访问）")
        if args.proxy or proxy_list:
            logger.info(f"使用代理: {args.proxy or f'{len(proxy_list)}个代理轮换'}")
        
        # 批量验证
        logger.info(f"开始验证 {len(keywords)} 个关键词...")
        results = validator.validate_batch(keywords)
        
        # 保存结果
        save_results_to_excel(args.input, results, args.output)
        
        # 统计结果
        yes_count = sum(1 for r in results if r.get('has_ads') == 'Yes')
        no_count = sum(1 for r in results if r.get('has_ads') == 'No')
        error_count = sum(1 for r in results if r.get('has_ads') == 'Error')
        
        logger.info("=" * 60)
        logger.info("验证完成！")
        logger.info(f"总计: {len(results)} 个关键词")
        logger.info(f"有广告: {yes_count} 个")
        logger.info(f"无广告: {no_count} 个")
        logger.info(f"错误: {error_count} 个")
        logger.info("=" * 60)
        
    except KeyboardInterrupt:
        logger.info("\n用户中断，程序退出")
    except Exception as e:
        logger.error(f"程序执行出错: {str(e)}", exc_info=True)


if __name__ == "__main__":
    main()

